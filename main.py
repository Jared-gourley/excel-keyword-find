import os
from openpyxl import load_workbook
from tkinter import Tk, Label, Entry, Button


# Searches a given workbook using its absolute path for a keyword
def searchWorkbook(workbookPath, workbookName, keyword):
    workbook = load_workbook(workbookPath)

    for sheet in workbook:
        for col in sheet:
            for cell in col:
                # If the keyword is found anywhere in the cell, complete match
                # Non case sensitive
                if keyword.lower() in str(cell.value).lower():
                    print('ITEM FOUND: "' + str(cell.value) + '" IN WORKBOOK "' +
                          workbookName + '" IN SHEET "' + sheet.title +
                          '" AT CELL "' + cell.coordinate + '"\n')

    workbook.close()


# Takes the data from the entry boxes and starts the search
def startSearch():
    folderPath = response1.get()
    keyword = response2.get()

    # Tries to open the given directory
    try:
        workbookNames = os.listdir(folderPath)
    except:
        print('Could not find folder given.')
        return

    # Runs a search for every excel sheet in the directory
    for workbookName in workbookNames:
        workbookPath = folderPath + '/' + workbookName
        searchWorkbook(workbookPath, workbookName, keyword)
    print('\n\nDONE SEARCHING\n\n')

# ------------------------- Tkinter UI code -------------------------
window = Tk()
window.title('Item Finder')
window.geometry("325x425")
window.resizable(False, False)

titleLabel = Label(window, text='Item Finder', font=('Arial', 32))
titleLabel.pack()

label1 = Label(window, text='Path to directory containing spreadsheets:', font=('Arial', 12))
label1.pack(pady=(10, 0))

response1 = Entry(window, font=('Arial', 12))
response1.insert(0, 'C:/example-folder')
response1.pack()

label2 = Label(window, text='Keyword to search for:', font=('Arial', 12))
label2.pack(pady=(20, 0))

response2 = Entry(window, font=('Arial', 12))
#response2.insert(0, 'Keyword here')
response2.pack()

label3 = Label(window, text='Make sure folder contains \nonly .xlsx spreadsheet files', font=('Arial', 13))
label3.pack(pady=(50, 0))


startButton = Button(text='Go', font=('Arial', 36), command=startSearch)
startButton.pack(pady=30)

def enterKeyPressed(e): # e for keyboard event handler (required for syntax)
    startSearch()

window.bind('<Return>', enterKeyPressed) # parameter e is in the method because of this
window.mainloop()
